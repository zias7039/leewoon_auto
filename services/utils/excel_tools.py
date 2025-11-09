# invoicegen/utils/excel_tools.py
# -*- coding: utf-8 -*-
import io
import streamlit as st
from openpyxl import load_workbook

def load_wb_and_guess_sheet(xlsx_file, target_sheet: str, show_warning: bool = True):
    try:
        wb_tmp = load_workbook(filename=io.BytesIO(xlsx_file.getvalue()), data_only=True)
        idx = wb_tmp.sheetnames.index(target_sheet) if target_sheet in wb_tmp.sheetnames else 0
        return st.selectbox("사용할 시트", wb_tmp.sheetnames, index=idx)
    except Exception:
        if show_warning:
            st.warning("엑셀 미리보기 중 문제가 발생했습니다. 생성 시도는 가능합니다.")
        return None
