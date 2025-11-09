# invoicegen/services/generator.py
# -*- coding: utf-8 -*-
import io
from dataclasses import dataclass
from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl import load_workbook
from docx import Document

from utils.docx_tools import make_replacer, replace_everywhere, collect_leftover_tokens
from utils.pdf_convert import convert_docx_to_pdf_bytes
from utils.paths import ensure_docx, ensure_pdf
from constants import TARGET_SHEET

@dataclass
class GenerateResult:
    out_name: str
    docx_bytes: bytes
    pdf_bytes: bytes | None
    pdf_ok: bool
    zip_bytes: bytes
    leftovers: list[str]

def generate_documents(xlsx_bytes: bytes, docx_tpl_bytes: bytes, sheet_name: str | None, out_name: str) -> GenerateResult:
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else (
        wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    )

    doc = Document(io.BytesIO(docx_tpl_bytes))
    replacer = make_replacer(ws)
    replace_everywhere(doc, replacer)

    buf = io.BytesIO()
    doc.save(buf)
    buf.seek(0)
    docx_bytes = buf.getvalue()

    pdf_bytes = convert_docx_to_pdf_bytes(docx_bytes)
    pdf_ok = pdf_bytes is not None

    doc_after = Document(io.BytesIO(docx_bytes))
    leftovers = collect_leftover_tokens(doc_after)

    zip_buf = io.BytesIO()
    with ZipFile(zip_buf, "w", ZIP_DEFLATED) as zf:
        zf.writestr(ensure_docx(out_name), docx_bytes)
        if pdf_ok:
            zf.writestr(ensure_pdf(out_name), pdf_bytes)
    zip_buf.seek(0)

    return GenerateResult(
        out_name=out_name.strip() or ensure_docx(out_name),
        docx_bytes=docx_bytes,
        pdf_bytes=pdf_bytes,
        pdf_ok=pdf_ok,
        zip_bytes=zip_buf.getvalue(),
        leftovers=leftovers,
    )
